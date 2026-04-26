import os
import re
import glob
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint

# ── DATABASE UNIVERSITAS ───────────────────────────────────────────────────────
# Format: { alias/keyword : (nama_lengkap, kota, provinsi, status) }
DAFTAR_UNIVERSITAS = {

    # ════════════════════════════════════════════════════════
    # DKI JAKARTA & SEKITARNYA
    # ════════════════════════════════════════════════════════
    "universitas indonesia":        ("Universitas Indonesia",            "Depok",             "Jawa Barat",        "Negeri"),
    " ui ":                         ("Universitas Indonesia",            "Depok",             "Jawa Barat",        "Negeri"),
    "fhui":                         ("Universitas Indonesia",            "Depok",             "Jawa Barat",        "Negeri"),
    "ftui":                         ("Universitas Indonesia",            "Depok",             "Jawa Barat",        "Negeri"),
    "fkui":                         ("Universitas Indonesia",            "Depok",             "Jawa Barat",        "Negeri"),
    "fisip ui":                     ("Universitas Indonesia",            "Depok",             "Jawa Barat",        "Negeri"),
    "unj":                          ("Universitas Negeri Jakarta",       "Jakarta",           "DKI Jakarta",       "Negeri"),
    "universitas negeri jakarta":   ("Universitas Negeri Jakarta",       "Jakarta",           "DKI Jakarta",       "Negeri"),
    "universitas trisakti":         ("Universitas Trisakti",             "Jakarta",           "DKI Jakarta",       "Swasta"),
    "trisakti":                     ("Universitas Trisakti",             "Jakarta",           "DKI Jakarta",       "Swasta"),
    "universitas tarumanagara":     ("Universitas Tarumanagara",         "Jakarta",           "DKI Jakarta",       "Swasta"),
    "tarumanagara":                 ("Universitas Tarumanagara",         "Jakarta",           "DKI Jakarta",       "Swasta"),
    "untar":                        ("Universitas Tarumanagara",         "Jakarta",           "DKI Jakarta",       "Swasta"),
    "universitas pancasila":        ("Universitas Pancasila",            "Jakarta",           "DKI Jakarta",       "Swasta"),
    "universitas bina nusantara":   ("Universitas Bina Nusantara",       "Jakarta",           "DKI Jakarta",       "Swasta"),
    "binus":                        ("Universitas Bina Nusantara",       "Jakarta",           "DKI Jakarta",       "Swasta"),
    "universitas atmajaya":         ("Universitas Atma Jaya Jakarta",    "Jakarta",           "DKI Jakarta",       "Swasta"),
    "atma jaya jakarta":            ("Universitas Atma Jaya Jakarta",    "Jakarta",           "DKI Jakarta",       "Swasta"),
    "uaj":                          ("Universitas Atma Jaya Jakarta",    "Jakarta",           "DKI Jakarta",       "Swasta"),
    "uin syarif hidayatullah":      ("UIN Syarif Hidayatullah",          "Tangerang Selatan", "Banten",            "Negeri"),
    "uin jakarta":                  ("UIN Syarif Hidayatullah",          "Tangerang Selatan", "Banten",            "Negeri"),
    "uinjkt":                       ("UIN Syarif Hidayatullah",          "Tangerang Selatan", "Banten",            "Negeri"),
    "universitas pertamina":        ("Universitas Pertamina",            "Jakarta",           "DKI Jakarta",       "Swasta"),
    "universitas gunadarma":        ("Universitas Gunadarma",            "Depok",             "Jawa Barat",        "Swasta"),
    "gunadarma":                    ("Universitas Gunadarma",            "Depok",             "Jawa Barat",        "Swasta"),
    "universitas mercu buana":      ("Universitas Mercu Buana",          "Jakarta",           "DKI Jakarta",       "Swasta"),
    "mercu buana":                  ("Universitas Mercu Buana",          "Jakarta",           "DKI Jakarta",       "Swasta"),
    "universitas pelita harapan":   ("Universitas Pelita Harapan",       "Tangerang",         "Banten",            "Swasta"),
    "uph":                          ("Universitas Pelita Harapan",       "Tangerang",         "Banten",            "Swasta"),
    "universitas pembangunan jaya": ("Universitas Pembangunan Jaya",     "Tangerang Selatan", "Banten",            "Swasta"),
    "universitas paramadina":       ("Universitas Paramadina",           "Jakarta",           "DKI Jakarta",       "Swasta"),
    "paramadina":                   ("Universitas Paramadina",           "Jakarta",           "DKI Jakarta",       "Swasta"),
    "universitas yarsi":            ("Universitas YARSI",                "Jakarta",           "DKI Jakarta",       "Swasta"),
    "yarsi":                        ("Universitas YARSI",                "Jakarta",           "DKI Jakarta",       "Swasta"),
    "universitas esa unggul":       ("Universitas Esa Unggul",           "Jakarta",           "DKI Jakarta",       "Swasta"),
    "esa unggul":                   ("Universitas Esa Unggul",           "Jakarta",           "DKI Jakarta",       "Swasta"),
    "pn jakarta":                   ("Politeknik Negeri Jakarta",        "Depok",             "Jawa Barat",        "Negeri"),
    "politeknik negeri jakarta":    ("Politeknik Negeri Jakarta",        "Depok",             "Jawa Barat",        "Negeri"),

    # ════════════════════════════════════════════════════════
    # JAWA BARAT (non-Depok)
    # ════════════════════════════════════════════════════════
    "itb":                          ("Institut Teknologi Bandung",       "Bandung",           "Jawa Barat",        "Negeri"),
    "institut teknologi bandung":   ("Institut Teknologi Bandung",       "Bandung",           "Jawa Barat",        "Negeri"),
    "unpad":                        ("Universitas Padjadjaran",          "Sumedang",          "Jawa Barat",        "Negeri"),
    "universitas padjadjaran":      ("Universitas Padjadjaran",          "Sumedang",          "Jawa Barat",        "Negeri"),
    "upi bandung":                  ("Universitas Pendidikan Indonesia", "Bandung",           "Jawa Barat",        "Negeri"),
    "universitas pendidikan indonesia": ("Universitas Pendidikan Indonesia", "Bandung",       "Jawa Barat",        "Negeri"),
    "telkom university":            ("Telkom University",                "Bandung",           "Jawa Barat",        "Swasta"),
    "universitas telkom":           ("Telkom University",                "Bandung",           "Jawa Barat",        "Swasta"),
    "universitas kristen maranatha": ("Universitas Kristen Maranatha",  "Bandung",           "Jawa Barat",        "Swasta"),
    "maranatha bandung":            ("Universitas Kristen Maranatha",   "Bandung",           "Jawa Barat",        "Swasta"),
    "universitas pasundan":         ("Universitas Pasundan",             "Bandung",           "Jawa Barat",        "Swasta"),
    "unpas":                        ("Universitas Pasundan",             "Bandung",           "Jawa Barat",        "Swasta"),
    "uin sunan gunung djati":       ("UIN Sunan Gunung Djati",           "Bandung",           "Jawa Barat",        "Negeri"),
    "uin bandung":                  ("UIN Sunan Gunung Djati",           "Bandung",           "Jawa Barat",        "Negeri"),
    "universitas komputer indonesia": ("Universitas Komputer Indonesia", "Bandung",           "Jawa Barat",        "Swasta"),
    "unikom":                       ("Universitas Komputer Indonesia",   "Bandung",           "Jawa Barat",        "Swasta"),
    "universitas widyatama":        ("Universitas Widyatama",            "Bandung",           "Jawa Barat",        "Swasta"),
    "widyatama":                    ("Universitas Widyatama",            "Bandung",           "Jawa Barat",        "Swasta"),
    "stba yapari":                  ("STBA Yapari",                      "Bandung",           "Jawa Barat",        "Swasta"),
    "ipb":                          ("Institut Pertanian Bogor",         "Bogor",             "Jawa Barat",        "Negeri"),
    "ipb university":               ("Institut Pertanian Bogor",         "Bogor",             "Jawa Barat",        "Negeri"),
    "institut pertanian bogor":     ("Institut Pertanian Bogor",         "Bogor",             "Jawa Barat",        "Negeri"),

    # ════════════════════════════════════════════════════════
    # JAWA TENGAH & DIY
    # ════════════════════════════════════════════════════════
    "ugm":                          ("Universitas Gadjah Mada",          "Yogyakarta",        "DI Yogyakarta",     "Negeri"),
    "universitas gadjah mada":      ("Universitas Gadjah Mada",          "Yogyakarta",        "DI Yogyakarta",     "Negeri"),
    "uny":                          ("Universitas Negeri Yogyakarta",    "Yogyakarta",        "DI Yogyakarta",     "Negeri"),
    "universitas negeri yogyakarta": ("Universitas Negeri Yogyakarta",   "Yogyakarta",        "DI Yogyakarta",     "Negeri"),
    "uii":                          ("Universitas Islam Indonesia",       "Yogyakarta",        "DI Yogyakarta",     "Swasta"),
    "universitas islam indonesia":  ("Universitas Islam Indonesia",       "Yogyakarta",        "DI Yogyakarta",     "Swasta"),
    "uajy":                         ("Universitas Atma Jaya Yogyakarta", "Yogyakarta",        "DI Yogyakarta",     "Swasta"),
    "atma jaya yogyakarta":         ("Universitas Atma Jaya Yogyakarta", "Yogyakarta",        "DI Yogyakarta",     "Swasta"),
    "uin sunan kalijaga":           ("UIN Sunan Kalijaga",               "Yogyakarta",        "DI Yogyakarta",     "Negeri"),
    "uin yogyakarta":               ("UIN Sunan Kalijaga",               "Yogyakarta",        "DI Yogyakarta",     "Negeri"),
    "universitas sanata dharma":    ("Universitas Sanata Dharma",        "Yogyakarta",        "DI Yogyakarta",     "Swasta"),
    "sanata dharma":                ("Universitas Sanata Dharma",        "Yogyakarta",        "DI Yogyakarta",     "Swasta"),
    "universitas muhammadiyah yogyakarta": ("Universitas Muhammadiyah Yogyakarta", "Yogyakarta", "DI Yogyakarta", "Swasta"),
    "umy":                          ("Universitas Muhammadiyah Yogyakarta", "Yogyakarta",      "DI Yogyakarta",     "Swasta"),
    "universitas amikom":           ("Universitas AMIKOM Yogyakarta",    "Yogyakarta",        "DI Yogyakarta",     "Swasta"),
    "amikom":                       ("Universitas AMIKOM Yogyakarta",    "Yogyakarta",        "DI Yogyakarta",     "Swasta"),
    "undip":                        ("Universitas Diponegoro",           "Semarang",          "Jawa Tengah",       "Negeri"),
    "universitas diponegoro":       ("Universitas Diponegoro",           "Semarang",          "Jawa Tengah",       "Negeri"),
    "unnes":                        ("Universitas Negeri Semarang",      "Semarang",          "Jawa Tengah",       "Negeri"),
    "universitas negeri semarang":  ("Universitas Negeri Semarang",      "Semarang",          "Jawa Tengah",       "Negeri"),
    "unika soegijapranata":         ("Universitas Katolik Soegijapranata", "Semarang",        "Jawa Tengah",       "Swasta"),
    "soegijapranata":               ("Universitas Katolik Soegijapranata", "Semarang",        "Jawa Tengah",       "Swasta"),
    "uns":                          ("Universitas Sebelas Maret",        "Surakarta",         "Jawa Tengah",       "Negeri"),
    "universitas sebelas maret":    ("Universitas Sebelas Maret",        "Surakarta",         "Jawa Tengah",       "Negeri"),
    "universitas muhammadiyah surakarta": ("Universitas Muhammadiyah Surakarta", "Surakarta", "Jawa Tengah",      "Swasta"),
    "ums solo":                     ("Universitas Muhammadiyah Surakarta", "Surakarta",       "Jawa Tengah",       "Swasta"),
    "universitas dian nuswantoro":  ("Universitas Dian Nuswantoro",      "Semarang",          "Jawa Tengah",       "Swasta"),
    "udinus":                       ("Universitas Dian Nuswantoro",      "Semarang",          "Jawa Tengah",       "Swasta"),
    "unsoed":                       ("Universitas Jenderal Soedirman",   "Purwokerto",        "Jawa Tengah",       "Negeri"),
    "universitas jenderal soedirman": ("Universitas Jenderal Soedirman","Purwokerto",         "Jawa Tengah",       "Negeri"),

    # ════════════════════════════════════════════════════════
    # JAWA TIMUR
    # ════════════════════════════════════════════════════════
    "unair":                        ("Universitas Airlangga",            "Surabaya",          "Jawa Timur",        "Negeri"),
    "universitas airlangga":        ("Universitas Airlangga",            "Surabaya",          "Jawa Timur",        "Negeri"),
    "its":                          ("Institut Teknologi Sepuluh Nopember", "Surabaya",        "Jawa Timur",        "Negeri"),
    "institut teknologi sepuluh nopember": ("Institut Teknologi Sepuluh Nopember", "Surabaya","Jawa Timur",        "Negeri"),
    "unesa":                        ("Universitas Negeri Surabaya",      "Surabaya",          "Jawa Timur",        "Negeri"),
    "universitas negeri surabaya":  ("Universitas Negeri Surabaya",      "Surabaya",          "Jawa Timur",        "Negeri"),
    "ubaya":                        ("Universitas Surabaya",             "Surabaya",          "Jawa Timur",        "Swasta"),
    "universitas surabaya":         ("Universitas Surabaya",             "Surabaya",          "Jawa Timur",        "Swasta"),
    "petra":                        ("Universitas Kristen Petra",        "Surabaya",          "Jawa Timur",        "Swasta"),
    "universitas kristen petra":    ("Universitas Kristen Petra",        "Surabaya",          "Jawa Timur",        "Swasta"),
    "universitas ciputra":          ("Universitas Ciputra",              "Surabaya",          "Jawa Timur",        "Swasta"),
    "ub":                           ("Universitas Brawijaya",            "Malang",            "Jawa Timur",        "Negeri"),
    "universitas brawijaya":        ("Universitas Brawijaya",            "Malang",            "Jawa Timur",        "Negeri"),
    "um malang":                    ("Universitas Negeri Malang",        "Malang",            "Jawa Timur",        "Negeri"),
    "universitas negeri malang":    ("Universitas Negeri Malang",        "Malang",            "Jawa Timur",        "Negeri"),
    "universitas muhammadiyah malang": ("Universitas Muhammadiyah Malang", "Malang",          "Jawa Timur",        "Swasta"),
    "umm":                          ("Universitas Muhammadiyah Malang",  "Malang",            "Jawa Timur",        "Swasta"),
    "universitas islam malang":     ("Universitas Islam Malang",         "Malang",            "Jawa Timur",        "Swasta"),
    "unisma":                       ("Universitas Islam Malang",         "Malang",            "Jawa Timur",        "Swasta"),
    "unej":                         ("Universitas Jember",               "Jember",            "Jawa Timur",        "Negeri"),
    "universitas jember":           ("Universitas Jember",               "Jember",            "Jawa Timur",        "Negeri"),
    "unirow":                       ("Universitas PGRI Ronggolawe",      "Tuban",             "Jawa Timur",        "Swasta"),
    "uin sunan ampel":              ("UIN Sunan Ampel",                  "Surabaya",          "Jawa Timur",        "Negeri"),
    "uinsa":                        ("UIN Sunan Ampel",                  "Surabaya",          "Jawa Timur",        "Negeri"),
    "universitas widyagama":        ("Universitas Widyagama",            "Malang",            "Jawa Timur",        "Swasta"),

    # ════════════════════════════════════════════════════════
    # SUMATERA UTARA
    # ════════════════════════════════════════════════════════
    "usu":                          ("Universitas Sumatera Utara",       "Medan",             "Sumatera Utara",    "Negeri"),
    "universitas sumatera utara":   ("Universitas Sumatera Utara",       "Medan",             "Sumatera Utara",    "Negeri"),
    "unimed":                       ("Universitas Negeri Medan",         "Medan",             "Sumatera Utara",    "Negeri"),
    "universitas negeri medan":     ("Universitas Negeri Medan",         "Medan",             "Sumatera Utara",    "Negeri"),
    "uin sumatera utara":           ("UIN Sumatera Utara",               "Medan",             "Sumatera Utara",    "Negeri"),
    "uinsu":                        ("UIN Sumatera Utara",               "Medan",             "Sumatera Utara",    "Negeri"),
    "universitas hkbp nommensen":   ("Universitas HKBP Nommensen",       "Medan",             "Sumatera Utara",    "Swasta"),
    "nommensen":                    ("Universitas HKBP Nommensen",       "Medan",             "Sumatera Utara",    "Swasta"),

    # ════════════════════════════════════════════════════════
    # SUMATERA BARAT
    # ════════════════════════════════════════════════════════
    "unand":                        ("Universitas Andalas",              "Padang",            "Sumatera Barat",    "Negeri"),
    "universitas andalas":          ("Universitas Andalas",              "Padang",            "Sumatera Barat",    "Negeri"),
    "unp padang":                   ("Universitas Negeri Padang",        "Padang",            "Sumatera Barat",    "Negeri"),
    "universitas negeri padang":    ("Universitas Negeri Padang",        "Padang",            "Sumatera Barat",    "Negeri"),
    "uin imam bonjol":              ("UIN Imam Bonjol",                  "Padang",            "Sumatera Barat",    "Negeri"),
    "iain imam bonjol":             ("UIN Imam Bonjol",                  "Padang",            "Sumatera Barat",    "Negeri"),

    # ════════════════════════════════════════════════════════
    # SUMATERA SELATAN
    # ════════════════════════════════════════════════════════
    "unsri":                        ("Universitas Sriwijaya",            "Palembang",         "Sumatera Selatan",  "Negeri"),
    "universitas sriwijaya":        ("Universitas Sriwijaya",            "Palembang",         "Sumatera Selatan",  "Negeri"),
    "uin raden fatah":              ("UIN Raden Fatah",                  "Palembang",         "Sumatera Selatan",  "Negeri"),

    # ════════════════════════════════════════════════════════
    # RIAU & KEPULAUAN RIAU
    # ════════════════════════════════════════════════════════
    "universitas riau":             ("Universitas Riau",                 "Pekanbaru",         "Riau",              "Negeri"),
    "unri":                         ("Universitas Riau",                 "Pekanbaru",         "Riau",              "Negeri"),
    "uin sultan syarif kasim":      ("UIN Sultan Syarif Kasim",          "Pekanbaru",         "Riau",              "Negeri"),
    "uin riau":                     ("UIN Sultan Syarif Kasim",          "Pekanbaru",         "Riau",              "Negeri"),
    "universitas maritim raja ali haji": ("Universitas Maritim Raja Ali Haji", "Tanjungpinang","Kepulauan Riau",  "Negeri"),
    "umrah":                        ("Universitas Maritim Raja Ali Haji","Tanjungpinang",      "Kepulauan Riau",   "Negeri"),

    # ════════════════════════════════════════════════════════
    # JAMBI & BENGKULU
    # ════════════════════════════════════════════════════════
    "universitas jambi":            ("Universitas Jambi",                "Jambi",             "Jambi",             "Negeri"),
    "unja":                         ("Universitas Jambi",                "Jambi",             "Jambi",             "Negeri"),
    "universitas bengkulu":         ("Universitas Bengkulu",             "Bengkulu",          "Bengkulu",          "Negeri"),
    "unib":                         ("Universitas Bengkulu",             "Bengkulu",          "Bengkulu",          "Negeri"),

    # ════════════════════════════════════════════════════════
    # LAMPUNG
    # ════════════════════════════════════════════════════════
    "unila":                        ("Universitas Lampung",              "Bandar Lampung",    "Lampung",           "Negeri"),
    "universitas lampung":          ("Universitas Lampung",              "Bandar Lampung",    "Lampung",           "Negeri"),

    # ════════════════════════════════════════════════════════
    # ACEH
    # ════════════════════════════════════════════════════════
    "unsyiah":                      ("Universitas Syiah Kuala",          "Banda Aceh",        "Aceh",              "Negeri"),
    "universitas syiah kuala":      ("Universitas Syiah Kuala",          "Banda Aceh",        "Aceh",              "Negeri"),
    "uin ar-raniry":                ("UIN Ar-Raniry",                    "Banda Aceh",        "Aceh",              "Negeri"),
    "uin aceh":                     ("UIN Ar-Raniry",                    "Banda Aceh",        "Aceh",              "Negeri"),

    # ════════════════════════════════════════════════════════
    # KALIMANTAN
    # ════════════════════════════════════════════════════════
    "ulm":                          ("Universitas Lambung Mangkurat",    "Banjarmasin",       "Kalimantan Selatan","Negeri"),
    "universitas lambung mangkurat": ("Universitas Lambung Mangkurat",   "Banjarmasin",       "Kalimantan Selatan","Negeri"),
    "untan":                        ("Universitas Tanjungpura",          "Pontianak",         "Kalimantan Barat",  "Negeri"),
    "universitas tanjungpura":      ("Universitas Tanjungpura",          "Pontianak",         "Kalimantan Barat",  "Negeri"),
    "unmul":                        ("Universitas Mulawarman",           "Samarinda",         "Kalimantan Timur",  "Negeri"),
    "universitas mulawarman":       ("Universitas Mulawarman",           "Samarinda",         "Kalimantan Timur",  "Negeri"),
    "universitas palangkaraya":     ("Universitas Palangka Raya",        "Palangka Raya",     "Kalimantan Tengah", "Negeri"),
    "upr":                          ("Universitas Palangka Raya",        "Palangka Raya",     "Kalimantan Tengah", "Negeri"),

    # ════════════════════════════════════════════════════════
    # SULAWESI
    # ════════════════════════════════════════════════════════
    "unhas":                        ("Universitas Hasanuddin",           "Makassar",          "Sulawesi Selatan",  "Negeri"),
    "universitas hasanuddin":       ("Universitas Hasanuddin",           "Makassar",          "Sulawesi Selatan",  "Negeri"),
    "unm makassar":                 ("Universitas Negeri Makassar",      "Makassar",          "Sulawesi Selatan",  "Negeri"),
    "universitas negeri makassar":  ("Universitas Negeri Makassar",      "Makassar",          "Sulawesi Selatan",  "Negeri"),
    "universitas muhammadiyah makassar": ("Universitas Muhammadiyah Makassar", "Makassar",    "Sulawesi Selatan",  "Swasta"),
    "unismuh":                      ("Universitas Muhammadiyah Makassar","Makassar",           "Sulawesi Selatan",  "Swasta"),
    "uin alauddin":                 ("UIN Alauddin",                     "Makassar",          "Sulawesi Selatan",  "Negeri"),
    "uin makassar":                 ("UIN Alauddin",                     "Makassar",          "Sulawesi Selatan",  "Negeri"),
    "unsrat":                       ("Universitas Sam Ratulangi",        "Manado",            "Sulawesi Utara",    "Negeri"),
    "universitas sam ratulangi":    ("Universitas Sam Ratulangi",        "Manado",            "Sulawesi Utara",    "Negeri"),
    "untad":                        ("Universitas Tadulako",             "Palu",              "Sulawesi Tengah",   "Negeri"),
    "universitas tadulako":         ("Universitas Tadulako",             "Palu",              "Sulawesi Tengah",   "Negeri"),
    "uin datokarama":               ("UIN Datokarama",                   "Palu",              "Sulawesi Tengah",   "Negeri"),
    "universitas halu oleo":        ("Universitas Halu Oleo",            "Kendari",           "Sulawesi Tenggara", "Negeri"),
    "uho":                          ("Universitas Halu Oleo",            "Kendari",           "Sulawesi Tenggara", "Negeri"),

    # ════════════════════════════════════════════════════════
    # BALI & NUSA TENGGARA
    # ════════════════════════════════════════════════════════
    "udayana":                      ("Universitas Udayana",              "Denpasar",          "Bali",              "Negeri"),
    "universitas udayana":          ("Universitas Udayana",              "Denpasar",          "Bali",              "Negeri"),
    "undiksha":                     ("Universitas Pendidikan Ganesha",   "Singaraja",         "Bali",              "Negeri"),
    "universitas pendidikan ganesha": ("Universitas Pendidikan Ganesha", "Singaraja",         "Bali",              "Negeri"),
    "stikes buleleng":              ("STIKES Buleleng",                  "Buleleng",          "Bali",              "Swasta"),
    "universitas mataram":          ("Universitas Mataram",              "Mataram",           "Nusa Tenggara Barat","Negeri"),
    "unram":                        ("Universitas Mataram",              "Mataram",           "Nusa Tenggara Barat","Negeri"),
    "universitas nusa cendana":     ("Universitas Nusa Cendana",         "Kupang",            "Nusa Tenggara Timur","Negeri"),
    "undana":                       ("Universitas Nusa Cendana",         "Kupang",            "Nusa Tenggara Timur","Negeri"),

    # ════════════════════════════════════════════════════════
    # MALUKU & PAPUA
    # ════════════════════════════════════════════════════════
    "universitas pattimura":        ("Universitas Pattimura",            "Ambon",             "Maluku",            "Negeri"),
    "unpatti":                      ("Universitas Pattimura",            "Ambon",             "Maluku",            "Negeri"),
    "universitas cenderawasih":     ("Universitas Cenderawasih",         "Jayapura",          "Papua",             "Negeri"),
    "uncen":                        ("Universitas Cenderawasih",         "Jayapura",          "Papua",             "Negeri"),

    # ════════════════════════════════════════════════════════
    # MUHAMMADIYAH LAINNYA
    # ════════════════════════════════════════════════════════
    "universitas muhammadiyah ponorogo": ("Universitas Muhammadiyah Ponorogo", "Ponorogo",    "Jawa Timur",        "Swasta"),
    "umpo":                         ("Universitas Muhammadiyah Ponorogo","Ponorogo",           "Jawa Timur",        "Swasta"),
    "universitas muhammadiyah palembang": ("Universitas Muhammadiyah Palembang", "Palembang", "Sumatera Selatan",  "Swasta"),
    "ump palembang":                ("Universitas Muhammadiyah Palembang","Palembang",         "Sumatera Selatan",  "Swasta"),

    # ════════════════════════════════════════════════════════
    # UMUM / TIDAK TERIDENTIFIKASI KOTA
    # ════════════════════════════════════════════════════════
    "iain":                         ("IAIN (tidak teridentifikasi)",     "Tidak Diketahui",   "Tidak Diketahui",   "Negeri"),
    "stikes":                       ("STIKES (tidak teridentifikasi)",   "Tidak Diketahui",   "Tidak Diketahui",   "Swasta"),
    "politeknik":                   ("Politeknik (tidak teridentifikasi)","Tidak Diketahui",  "Tidak Diketahui",   "Negeri"),
}

# ── JENIS KEKERASAN ────────────────────────────────────────────────────────────
JENIS_KEKERASAN = {
    "pemerkosaan":              "Pemerkosaan",
    "perkosaan":                "Pemerkosaan",
    "sexual assault":           "Pemerkosaan",
    "diperkosa":                "Pemerkosaan",
    "pelecehan seksual":        "Pelecehan Seksual",
    "pelecehan":                "Pelecehan Seksual",
    "catcalling":               "Pelecehan Seksual",
    "sexual harassment":        "Pelecehan Seksual",
    "kekerasan seksual":        "Kekerasan Seksual (Umum)",
    "sexual violence":          "Kekerasan Seksual (Umum)",
    "kekerasan berbasis gender": "KBGO",
    "kbgo":                     "KBGO",
    "eksploitasi seksual":      "Eksploitasi Seksual",
    "konten seksual":           "Penyebaran Konten Seksual",
    "revenge porn":             "Penyebaran Konten Seksual",
    "foto bugil":               "Penyebaran Konten Seksual",
    "video asusila":            "Penyebaran Konten Seksual",
    "grooming":                 "Grooming",
    "pelecehan verbal":         "Pelecehan Verbal",
    "intimidasi seksual":       "Intimidasi/Ancaman Seksual",
    "ancaman seksual":          "Intimidasi/Ancaman Seksual",
    "pemaksaan":                "Pemaksaan Seksual",
}

# ── PELAKU ─────────────────────────────────────────────────────────────────────
PELAKU = {
    "dosen pembimbing":         "Dosen Pembimbing",
    "dosen":                    "Dosen",
    "profesor":                 "Dosen",
    "guru besar":               "Dosen",
    "rektor":                   "Rektor/Pimpinan",
    "dekan":                    "Rektor/Pimpinan",
    "wakil rektor":             "Rektor/Pimpinan",
    "direktur":                 "Rektor/Pimpinan",
    "ketua program studi":      "Rektor/Pimpinan",
    "kaprodi":                  "Rektor/Pimpinan",
    "mahasiswa":                "Mahasiswa",
    "senior":                   "Senior/Kakak Tingkat",
    "kakak tingkat":            "Senior/Kakak Tingkat",
    "bem":                      "Pengurus BEM/Organisasi",
    "ketua bem":                "Pengurus BEM/Organisasi",
    "ketua organisasi":         "Pengurus BEM/Organisasi",
    "panitia":                  "Panitia Kegiatan",
    "tendik":                   "Tenaga Kependidikan",
    "satpam":                   "Tenaga Kependidikan",
    "karyawan":                 "Tenaga Kependidikan",
    "alumni":                   "Alumni",
}

# ── EKSTRAKSI ──────────────────────────────────────────────────────────────────
def ekstrak_universitas(teks: str):
    """Deteksi universitas dari teks. Return (nama, kota, provinsi, status)."""
    if not teks:
        return None, None, None, None
    teks_lower = " " + teks.lower() + " "
    for alias, info in sorted(DAFTAR_UNIVERSITAS.items(), key=lambda x: -len(x[0])):
        if alias in teks_lower:
            return info  # (nama, kota, provinsi, status)
    return None, None, None, None


def ekstrak_jenis_kekerasan(teks: str):
    if not teks:
        return "Tidak Teridentifikasi"
    t = teks.lower()
    for kw, label in sorted(JENIS_KEKERASAN.items(), key=lambda x: -len(x[0])):
        if kw in t:
            return label
    return "Tidak Teridentifikasi"


def ekstrak_pelaku(teks: str):
    if not teks:
        return "Tidak Diketahui"
    t = teks.lower()
    for kw, label in sorted(PELAKU.items(), key=lambda x: -len(x[0])):
        if kw in t:
            return label
    return "Tidak Diketahui"


def proses_satu_file(path_csv: str) -> pd.DataFrame:
    try:
        df = pd.read_csv(path_csv, encoding="utf-8-sig")
    except UnicodeDecodeError:
        df = pd.read_csv(path_csv, encoding="cp1252")

    col_teks = [c for c in ["judul", "title", "deskripsi", "description",
                             "konten", "content", "komentar", "text"] if c in df.columns]

    def gabung_teks(row):
        return " ".join(str(row[c]) for c in col_teks if pd.notna(row.get(c, "")))

    df["_teks"] = df.apply(gabung_teks, axis=1)

    ekstrak = df["_teks"].apply(
        lambda t: pd.Series(
            ekstrak_universitas(t),
            index=["universitas", "kota", "provinsi", "status_pt"]
        )
    )
    df["universitas"]     = ekstrak["universitas"].fillna("Tidak Teridentifikasi")
    df["kota"]            = ekstrak["kota"].fillna("Tidak Diketahui")
    df["provinsi"]        = ekstrak["provinsi"].fillna("Tidak Diketahui")
    df["status_pt"]       = ekstrak["status_pt"].fillna("Tidak Diketahui")
    df["jenis_kekerasan"] = df["_teks"].apply(ekstrak_jenis_kekerasan)
    df["pelaku"]          = df["_teks"].apply(ekstrak_pelaku)

    col_tgl = next((c for c in ["tanggal", "tanggal_tayang", "date", "publishedAt"]
                    if c in df.columns), None)
    if col_tgl:
        tgl = pd.to_datetime(df[col_tgl], errors="coerce", utc=True)
        df["tahun"] = tgl.dt.year.astype("Int64")
        df["bulan"] = tgl.dt.month_name()
        df["bulan_num"] = tgl.dt.month.astype("Int64")
    else:
        df["tahun"] = pd.NA
        df["bulan"] = pd.NA
        df["bulan_num"] = pd.NA

    df["_sumber_file"] = os.path.basename(path_csv)
    df.drop(columns=["_teks"], inplace=True)
    return df


# ── STYLING HELPERS ────────────────────────────────────────────────────────────
WARNA = {
    "biru_tua":     "1F497D",
    "hijau_tua":    "375623",
    "abu":          "595959",
    "merah":        "C00000",
    "ungu":         "7030A0",
    "oranye":       "C55A11",
    "row_hijau":    "E2EFDA",
    "row_biru":     "DCE6F1",
    "row_merah":    "FCE4D6",
    "kuning":       "FFF2CC",
    "putih":        "FFFFFF",
}


def style_header(ws, row_num, fill_hex, cols, font_size=10):
    thin = Side(style="thin", color="FFFFFF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = PatternFill("solid", fgColor=fill_hex)
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=font_size)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def set_col_widths(ws, mapping: dict, df_cols, default=15):
    for i, col_name in enumerate(df_cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = mapping.get(col_name, default)


def warnai_baris(ws, df, start_row=2, fill_col="universitas",
                 fill_yes=WARNA["row_hijau"], fill_no=WARNA["row_biru"]):
    fill_y = PatternFill("solid", fgColor=fill_yes)
    fill_n = PatternFill("solid", fgColor=fill_no)
    cols = len(df.columns)
    for r_idx, (_, row) in enumerate(df.iterrows(), start=start_row):
        val = str(row.get(fill_col, ""))
        fill = fill_y if val not in ("Tidak Teridentifikasi", "Tidak Diketahui", "") else (
               fill_n if r_idx % 2 == 0 else None)
        if fill:
            for c in range(1, cols + 1):
                ws.cell(row=r_idx, column=c).fill = fill


# ── BUAT EXCEL ─────────────────────────────────────────────────────────────────
def buat_excel_riset(df_all: pd.DataFrame, path_output: str):

    # ── Susun urutan kolom ──────────────────────────────────────────────────
    prioritas = [
        "universitas", "kota", "provinsi", "status_pt",
        "jenis_kekerasan", "pelaku", "tahun", "bulan",
        "judul", "sumber", "url", "tanggal", "tanggal_tayang",
        "views", "likes", "komentar", "keyword", "metode",
        "deskripsi", "konten", "_sumber_file",
    ]
    cols_ada   = [c for c in prioritas if c in df_all.columns]
    cols_sisa  = [c for c in df_all.columns if c not in cols_ada and c != "bulan_num"]
    df_out     = df_all[cols_ada + cols_sisa].copy()

    lebar_umum = {
        "universitas": 40, "kota": 22, "provinsi": 25, "status_pt": 12,
        "jenis_kekerasan": 28, "pelaku": 22, "tahun": 8, "bulan": 14,
        "judul": 60, "sumber": 18, "url": 45, "tanggal": 22,
        "tanggal_tayang": 22, "deskripsi": 65, "konten": 65,
        "keyword": 35, "_sumber_file": 30,
    }

    with pd.ExcelWriter(path_output, engine="openpyxl") as writer:

        # ════════════════════════════════════════════════════
        # SHEET 1 — DATA LENGKAP
        # ════════════════════════════════════════════════════
        df_out.to_excel(writer, sheet_name="Data Lengkap", index=False)
        ws1 = writer.sheets["Data Lengkap"]
        style_header(ws1, 1, WARNA["biru_tua"], len(df_out.columns))
        ws1.freeze_panes = "A2"
        ws1.auto_filter.ref = ws1.dimensions
        warnai_baris(ws1, df_out)
        set_col_widths(ws1, lebar_umum, df_out.columns)

        # ════════════════════════════════════════════════════
        # SHEET 2 — PER UNIVERSITAS (detail lengkap)
        # ════════════════════════════════════════════════════
        df_u = df_all[df_all["universitas"] != "Tidak Teridentifikasi"].copy()
        rekap_univ = (
            df_u.groupby(["universitas", "kota", "provinsi", "status_pt"])
            .agg(
                jumlah_kasus      = ("universitas", "count"),
                tahun_pertama     = ("tahun", "min"),
                tahun_terakhir    = ("tahun", "max"),
                jenis_kekerasan   = ("jenis_kekerasan",
                                     lambda x: " | ".join(sorted(set(x) - {"Tidak Teridentifikasi"}))),
                pelaku_terlibat   = ("pelaku",
                                     lambda x: " | ".join(sorted(set(x) - {"Tidak Diketahui"}))),
            )
            .reset_index()
            .sort_values("jumlah_kasus", ascending=False)
            .reset_index(drop=True)
        )
        rekap_univ.index = rekap_univ.index + 1   # ranking mulai 1
        rekap_univ.insert(0, "peringkat", rekap_univ.index)

        rekap_univ.to_excel(writer, sheet_name="Per Universitas", index=False)
        ws2 = writer.sheets["Per Universitas"]
        style_header(ws2, 1, WARNA["hijau_tua"], len(rekap_univ.columns))
        ws2.freeze_panes = "A2"
        ws2.auto_filter.ref = ws2.dimensions
        # Highlight 5 besar dengan merah
        fill_top = PatternFill("solid", fgColor=WARNA["row_merah"])
        fill_alt = PatternFill("solid", fgColor=WARNA["row_hijau"])
        for r_idx in range(2, len(rekap_univ) + 2):
            rank = ws2.cell(row=r_idx, column=1).value
            fill = fill_top if (rank and rank <= 5) else (fill_alt if r_idx % 2 == 0 else None)
            if fill:
                for c in range(1, len(rekap_univ.columns) + 1):
                    ws2.cell(row=r_idx, column=c).fill = fill
        set_col_widths(ws2, {
            "peringkat": 10, "universitas": 40, "kota": 22, "provinsi": 25,
            "status_pt": 12, "jumlah_kasus": 14, "tahun_pertama": 14,
            "tahun_terakhir": 14, "jenis_kekerasan": 45, "pelaku_terlibat": 45,
        }, rekap_univ.columns)
        # Tambah catatan merah = 5 besar
        ws2.cell(row=len(rekap_univ) + 3, column=1,
                 value="* Merah = 5 universitas dengan kasus terbanyak").font = \
            Font(name="Calibri", italic=True, color=WARNA["merah"], size=9)

        # ════════════════════════════════════════════════════
        # SHEET 3 — PER PROVINSI
        # ════════════════════════════════════════════════════
        rekap_prov = (
            df_all[df_all["provinsi"] != "Tidak Diketahui"]
            .groupby("provinsi")
            .agg(
                jumlah_kasus         = ("provinsi", "count"),
                universitas_terlibat = ("universitas",
                                        lambda x: len(set(x) - {"Tidak Teridentifikasi"})),
                nama_universitas     = ("universitas",
                                        lambda x: " | ".join(
                                            sorted(set(x) - {"Tidak Teridentifikasi"}))),
                negeri               = ("status_pt",
                                        lambda x: (x == "Negeri").sum()),
                swasta               = ("status_pt",
                                        lambda x: (x == "Swasta").sum()),
            )
            .reset_index()
            .sort_values("jumlah_kasus", ascending=False)
        )
        rekap_prov.to_excel(writer, sheet_name="Per Provinsi", index=False)
        ws3 = writer.sheets["Per Provinsi"]
        style_header(ws3, 1, WARNA["ungu"], len(rekap_prov.columns))
        ws3.freeze_panes = "A2"
        ws3.auto_filter.ref = ws3.dimensions
        warnai_baris(ws3, rekap_prov, fill_col="provinsi",
                     fill_yes=WARNA["row_biru"], fill_no=WARNA["row_hijau"])
        set_col_widths(ws3, {
            "provinsi": 28, "jumlah_kasus": 14, "universitas_terlibat": 22,
            "nama_universitas": 80, "negeri": 12, "swasta": 12,
        }, rekap_prov.columns)

        # ════════════════════════════════════════════════════
        # SHEET 4 — PER KOTA
        # ════════════════════════════════════════════════════
        rekap_kota = (
            df_all[df_all["kota"] != "Tidak Diketahui"]
            .groupby(["kota", "provinsi"])
            .agg(
                jumlah_kasus         = ("kota", "count"),
                universitas_terlibat = ("universitas",
                                        lambda x: " | ".join(sorted(set(x) - {"Tidak Teridentifikasi"}))),
            )
            .reset_index()
            .sort_values("jumlah_kasus", ascending=False)
        )
        rekap_kota.to_excel(writer, sheet_name="Per Kota", index=False)
        ws4 = writer.sheets["Per Kota"]
        style_header(ws4, 1, WARNA["abu"], len(rekap_kota.columns))
        ws4.freeze_panes = "A2"
        ws4.auto_filter.ref = ws4.dimensions
        set_col_widths(ws4, {
            "kota": 25, "provinsi": 25, "jumlah_kasus": 14,
            "universitas_terlibat": 80,
        }, rekap_kota.columns)

        # ════════════════════════════════════════════════════
        # SHEET 5 — JENIS KEKERASAN
        # ════════════════════════════════════════════════════
        rekap_jk = (
            df_all.groupby("jenis_kekerasan")
            .agg(
                jumlah           = ("jenis_kekerasan", "count"),
                persen           = ("jenis_kekerasan",
                                    lambda x: f"{100 * len(x) / len(df_all):.1f}%"),
                universitas_unik = ("universitas",
                                    lambda x: len(set(x) - {"Tidak Teridentifikasi"})),
                tahun_pertama    = ("tahun", "min"),
                tahun_terakhir   = ("tahun", "max"),
            )
            .reset_index()
            .sort_values("jumlah", ascending=False)
        )
        rekap_jk.to_excel(writer, sheet_name="Jenis Kekerasan", index=False)
        ws5 = writer.sheets["Jenis Kekerasan"]
        style_header(ws5, 1, WARNA["merah"], len(rekap_jk.columns))
        ws5.freeze_panes = "A2"
        ws5.auto_filter.ref = ws5.dimensions
        set_col_widths(ws5, {
            "jenis_kekerasan": 35, "jumlah": 12, "persen": 10,
            "universitas_unik": 18, "tahun_pertama": 14, "tahun_terakhir": 14,
        }, rekap_jk.columns)

        # ════════════════════════════════════════════════════
        # SHEET 6 — PELAKU
        # ════════════════════════════════════════════════════
        rekap_pelaku = (
            df_all.groupby("pelaku")
            .agg(
                jumlah           = ("pelaku", "count"),
                persen           = ("pelaku",
                                    lambda x: f"{100 * len(x) / len(df_all):.1f}%"),
                universitas_unik = ("universitas",
                                    lambda x: len(set(x) - {"Tidak Teridentifikasi"})),
            )
            .reset_index()
            .sort_values("jumlah", ascending=False)
        )
        rekap_pelaku.to_excel(writer, sheet_name="Pelaku", index=False)
        ws6 = writer.sheets["Pelaku"]
        style_header(ws6, 1, WARNA["oranye"], len(rekap_pelaku.columns))
        ws6.freeze_panes = "A2"
        ws6.auto_filter.ref = ws6.dimensions
        set_col_widths(ws6, {
            "pelaku": 30, "jumlah": 12, "persen": 10, "universitas_unik": 18,
        }, rekap_pelaku.columns)

        # ════════════════════════════════════════════════════
        # SHEET 7 — TABEL SILANG: Kekerasan × Pelaku
        # ════════════════════════════════════════════════════
        df_cross = df_all[
            (df_all["jenis_kekerasan"] != "Tidak Teridentifikasi") &
            (df_all["pelaku"] != "Tidak Diketahui")
        ]
        if not df_cross.empty:
            pivot = pd.crosstab(
                df_cross["jenis_kekerasan"],
                df_cross["pelaku"],
                margins=True, margins_name="TOTAL"
            )
            pivot.to_excel(writer, sheet_name="Tabel Silang JK×Pelaku")
            ws7 = writer.sheets["Tabel Silang JK×Pelaku"]
            style_header(ws7, 1, WARNA["biru_tua"], len(pivot.columns) + 1)
            ws7.freeze_panes = "B2"
            ws7.column_dimensions["A"].width = 35
            for ci in range(2, len(pivot.columns) + 2):
                ws7.column_dimensions[get_column_letter(ci)].width = 18

        # ════════════════════════════════════════════════════
        # SHEET 8 — TREN PER TAHUN
        # ════════════════════════════════════════════════════
        if "tahun" in df_all.columns:
            rekap_tahun = (
                df_all.dropna(subset=["tahun"])
                .groupby("tahun")
                .agg(
                    jumlah_kasus        = ("tahun", "count"),
                    teridentifikasi_pt  = ("universitas",
                                           lambda x: (x != "Tidak Teridentifikasi").sum()),
                    universitas_unik    = ("universitas",
                                           lambda x: len(set(x) - {"Tidak Teridentifikasi"})),
                    provinsi_unik       = ("provinsi",
                                           lambda x: len(set(x) - {"Tidak Diketahui"})),
                )
                .reset_index()
                .sort_values("tahun")
            )
            rekap_tahun["tahun"] = rekap_tahun["tahun"].astype(int)
            rekap_tahun.to_excel(writer, sheet_name="Tren per Tahun", index=False)
            ws8 = writer.sheets["Tren per Tahun"]
            style_header(ws8, 1, WARNA["biru_tua"], len(rekap_tahun.columns))
            ws8.freeze_panes = "A2"
            for ci in range(1, len(rekap_tahun.columns) + 1):
                ws8.column_dimensions[get_column_letter(ci)].width = 22

        # ════════════════════════════════════════════════════
        # SHEET 9 — DASHBOARD RINGKASAN
        # ════════════════════════════════════════════════════
        ws9 = writer.book.create_sheet("Dashboard Ringkasan")
        total         = len(df_all)
        teridentif    = (df_all["universitas"] != "Tidak Teridentifikasi").sum()
        univ_unik     = df_all[df_all["universitas"] != "Tidak Teridentifikasi"]["universitas"].nunique()
        kota_unik     = df_all[df_all["kota"] != "Tidak Diketahui"]["kota"].nunique()
        prov_unik     = df_all[df_all["provinsi"] != "Tidak Diketahui"]["provinsi"].nunique()
        top5          = rekap_univ.head(5)[["universitas", "provinsi", "jumlah_kasus"]]
        top3_prov     = rekap_prov.head(3)[["provinsi", "jumlah_kasus"]]
        top3_jk       = rekap_jk.head(3)[["jenis_kekerasan", "jumlah"]]
        top3_pelaku   = rekap_pelaku.head(3)[["pelaku", "jumlah"]]

        dash = [
            ["DASHBOARD RINGKASAN RISET", ""],
            ["Riset Kekerasan Seksual — Pendidikan Tinggi Indonesia", ""],
            [f"Dibuat: {datetime.now().strftime('%d %B %Y, %H:%M')}", ""],
            ["", ""],
            ["STATISTIK UTAMA", ""],
            ["Total data/artikel", f"{total:,}"],
            ["Data teridentifikasi PT", f"{teridentif:,}  ({100*teridentif/total:.1f}%)"],
            ["Universitas unik terdeteksi", f"{univ_unik}"],
            ["Kota unik", f"{kota_unik}"],
            ["Provinsi unik", f"{prov_unik}"],
            ["", ""],
            ["5 UNIVERSITAS KASUS TERBANYAK", "Jumlah"],
        ]
        for _, row in top5.iterrows():
            dash.append([f"{row['universitas']} ({row['provinsi']})", row["jumlah_kasus"]])
        dash += [
            ["", ""],
            ["3 PROVINSI KASUS TERBANYAK", "Jumlah"],
        ]
        for _, row in top3_prov.iterrows():
            dash.append([row["provinsi"], row["jumlah_kasus"]])
        dash += [
            ["", ""],
            ["3 JENIS KEKERASAN TERBANYAK", "Jumlah"],
        ]
        for _, row in top3_jk.iterrows():
            dash.append([row["jenis_kekerasan"], row["jumlah"]])
        dash += [
            ["", ""],
            ["3 PELAKU TERBANYAK", "Jumlah"],
        ]
        for _, row in top3_pelaku.iterrows():
            dash.append([row["pelaku"], row["jumlah"]])
        dash += [
            ["", ""],
            ["SUMBER DATA RESMI", "URL"],
            ["SIMFONI PPA – Kemenppa", "https://kekerasan.kemenpppa.go.id"],
            ["CATAHU – Komnas Perempuan", "https://komnasperempuan.go.id"],
            ["Satgas PPKS – Kemendiktisaintek", "(ajukan permintaan data resmi)"],
        ]

        for r_idx, (a, b) in enumerate(dash, start=1):
            ca = ws9.cell(row=r_idx, column=1, value=a)
            cb = ws9.cell(row=r_idx, column=2, value=b)
            if r_idx in (1, 2):
                ca.fill = PatternFill("solid", fgColor=WARNA["biru_tua"])
                ca.font = Font(name="Calibri", bold=True, color="FFFFFF",
                               size=14 if r_idx == 1 else 11)
                cb.fill = PatternFill("solid", fgColor=WARNA["biru_tua"])
            elif r_idx == 3:
                ca.font = Font(name="Calibri", italic=True, color=WARNA["abu"], size=9)
            elif a in ("STATISTIK UTAMA", "5 UNIVERSITAS KASUS TERBANYAK",
                        "3 PROVINSI KASUS TERBANYAK", "3 JENIS KEKERASAN TERBANYAK",
                        "3 PELAKU TERBANYAK", "SUMBER DATA RESMI"):
                for cell in [ca, cb]:
                    cell.fill = PatternFill("solid", fgColor=WARNA["hijau_tua"])
                    cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
            elif a and b:
                ca.font = Font(name="Calibri", size=10)
                cb.font = Font(name="Calibri", bold=True, size=10)
                cb.alignment = Alignment(horizontal="center")
                if r_idx % 2 == 0:
                    ca.fill = PatternFill("solid", fgColor=WARNA["row_biru"])
                    cb.fill = PatternFill("solid", fgColor=WARNA["row_biru"])

        ws9.column_dimensions["A"].width = 52
        ws9.column_dimensions["B"].width = 20

        # ════════════════════════════════════════════════════
        # SHEET 10 — PANDUAN & LEGENDA
        # ════════════════════════════════════════════════════
        ws10 = writer.book.create_sheet("Panduan & Legenda")
        panduan = [
            ["PANDUAN MEMBACA DATA", ""],
            ["", ""],
            ["Kolom", "Keterangan"],
            ["universitas", "Nama universitas yang terdeteksi dari teks berita/komentar"],
            ["kota",        "Kota lokasi kampus universitas"],
            ["provinsi",    "Provinsi lokasi kampus universitas (BARU v2.0)"],
            ["status_pt",   "Negeri = PTN | Swasta = PTS"],
            ["jenis_kekerasan", "Jenis kekerasan yang disebut dalam teks"],
            ["pelaku",     "Peran pelaku yang disebut (dosen, mahasiswa, dll.)"],
            ["tahun / bulan", "Diambil dari kolom tanggal artikel/video"],
            ["", ""],
            ["WARNA BARIS — Data Lengkap", ""],
            ["Hijau muda",  "Universitas BERHASIL diidentifikasi"],
            ["Biru muda",   "Baris genap — universitas tidak teridentifikasi"],
            ["Putih",       "Baris ganjil — universitas tidak teridentifikasi"],
            ["", ""],
            ["WARNA BARIS — Per Universitas", ""],
            ["Merah muda",  "5 universitas dengan kasus TERBANYAK"],
            ["Hijau muda",  "Universitas lainnya (baris genap)"],
            ["", ""],
            ["SHEET BARU v2.0", ""],
            ["Per Provinsi",         "Rekap kasus per provinsi + jumlah PTN & PTS"],
            ["Tabel Silang JK×Pelaku", "Crosstab jenis kekerasan vs. pelaku"],
            ["Dashboard Ringkasan",  "Statistik cepat & top-5 dalam satu halaman"],
            ["", ""],
            ["CATATAN PENTING", ""],
            ["1", "Data dari berita publik & YouTube — BUKAN data resmi pemerintah"],
            ["2", "Satu artikel bisa melaporkan beberapa kasus sekaligus"],
            ["3", "Kasus yang tidak sebut nama PT tidak dapat diidentifikasi"],
            ["4", "Untuk sitasi akademis, verifikasi ke sumber asli"],
            ["", ""],
            ["SUMBER DATA RESMI", ""],
            ["SIMFONI PPA – Kemenppa",     "https://kekerasan.kemenpppa.go.id"],
            ["CATAHU – Komnas Perempuan",  "https://komnasperempuan.go.id"],
            ["Satgas PPKS",               "Ajukan permintaan data ke Kemendiktisaintek"],
        ]
        for r, (a, b) in enumerate(panduan, start=1):
            ca = ws10.cell(row=r, column=1, value=a)
            cb = ws10.cell(row=r, column=2, value=b)
            ca.font = Font(name="Calibri", size=10,
                           bold=(a in ["PANDUAN MEMBACA DATA", "Kolom",
                                        "WARNA BARIS — Data Lengkap",
                                        "WARNA BARIS — Per Universitas",
                                        "SHEET BARU v2.0",
                                        "CATATAN PENTING", "SUMBER DATA RESMI"]))
            cb.font = Font(name="Calibri", size=10)
        ws10.column_dimensions["A"].width = 32
        ws10.column_dimensions["B"].width = 72
        for r_idx, (a, _) in enumerate(panduan, start=1):
            if a in ("PANDUAN MEMBACA DATA", "Kolom",
                      "WARNA BARIS — Data Lengkap", "WARNA BARIS — Per Universitas",
                      "SHEET BARU v2.0", "CATATAN PENTING", "SUMBER DATA RESMI"):
                clr = WARNA["biru_tua"] if a == "PANDUAN MEMBACA DATA" else (
                      WARNA["hijau_tua"] if "WARNA" in a or "SHEET" in a else WARNA["abu"])
                for ci in [1, 2]:
                    cell = ws10.cell(row=r_idx, column=ci)
                    cell.fill = PatternFill("solid", fgColor=clr)
                    cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)

    print(f"  ✓ Excel tersimpan: {path_output}")


# ── MAIN ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 68)
    print("  EKSTRAK UNIVERSITAS & RAPIKAN DATA  [v2.0]")
    print("  Riset Kekerasan Seksual — Pendidikan Tinggi Indonesia")
    print(f"  Mulai: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print("=" * 68)

    pola      = os.path.join("data", "*.csv")
    semua_csv = glob.glob(pola)

    if not semua_csv:
        print("\n  [ERROR] Tidak ada file CSV di folder data/")
        print("  Jalankan dulu scraper berita atau YouTube.")
        input("\n  Tekan ENTER untuk keluar...")
        exit(1)

    print(f"\n  Ditemukan {len(semua_csv)} file CSV:")
    for f in semua_csv:
        print(f"    • {os.path.basename(f)}")

    semua_df = []
    for path in semua_csv:
        print(f"\n  Memproses: {os.path.basename(path)} ...")
        try:
            df = proses_satu_file(path)
            semua_df.append(df)
            n_total   = len(df)
            n_identif = (df["universitas"] != "Tidak Teridentifikasi").sum()
            print(f"  → {n_total:,} baris | "
                  f"{n_identif:,} teridentifikasi PT "
                  f"({100*n_identif/max(n_total,1):.1f}%)")
        except Exception as e:
            print(f"  [WARN] Gagal memproses: {e}")

    if not semua_df:
        print("\n  [ERROR] Tidak ada data berhasil dibaca.")
        input("\n  Tekan ENTER untuk keluar...")
        exit(1)

    df_all     = pd.concat(semua_df, ignore_index=True)
    total      = len(df_all)
    teridentif = (df_all["universitas"] != "Tidak Teridentifikasi").sum()
    univ_unik  = df_all[df_all["universitas"] != "Tidak Teridentifikasi"]["universitas"].nunique()
    kota_unik  = df_all[df_all["kota"] != "Tidak Diketahui"]["kota"].nunique()
    prov_unik  = df_all[df_all["provinsi"] != "Tidak Diketahui"]["provinsi"].nunique()

    print(f"\n{'=' * 68}")
    print("  RINGKASAN HASIL:")
    print(f"  Total data          : {total:,}")
    print(f"  Teridentifikasi PT  : {teridentif:,}  ({100*teridentif/total:.1f}%)")
    print(f"  Universitas unik    : {univ_unik}")
    print(f"  Kota unik           : {kota_unik}")
    print(f"  Provinsi unik       : {prov_unik}")

    ts        = datetime.now().strftime("%Y%m%d_%H%M")
    os.makedirs("data", exist_ok=True)
    path_xlsx = os.path.join("data", f"RISET_kekerasan_seksual_PT_{ts}.xlsx")

    print(f"\n  Menyusun Excel riset: {path_xlsx}")
    buat_excel_riset(df_all, path_xlsx)

    print(f"\n{'=' * 68}")
    print("  SELESAI! Sheet yang tersedia:")
    print("    1. Data Lengkap           — semua data + kolom ekstraksi")
    print("    2. Per Universitas         — ranking PT + detail jenis & pelaku")
    print("    3. Per Provinsi            — rekap per provinsi")
    print("    4. Per Kota               — rekap per kota")
    print("    5. Jenis Kekerasan        — statistik per jenis")
    print("    6. Pelaku                 — statistik per pelaku")
    print("    7. Tabel Silang JK×Pelaku — crosstab kekerasan vs. pelaku")
    print("    8. Tren per Tahun         — tren kasus per tahun")
    print("    9. Dashboard Ringkasan    — satu halaman statistik utama")
    print("   10. Panduan & Legenda      — kamus kolom & warna")
    print(f"{'=' * 68}")
    input("\n  Tekan ENTER untuk keluar...")